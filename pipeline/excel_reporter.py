"""
pipeline/excel_reporter.py
──────────────────────────
Generates / updates Candidate_Ranking.xlsx

Rules
─────
• If the file does not exist → create it fresh
• If the file exists:
    - For each role sheet that already exists → append only NEW candidates
      (deduplicated by file_hash so the same resume never appears twice)
    - For each role that has no sheet yet → create a new sheet
• Every run updates the Summary sheet with current totals
• Scores are stored as 0-100 integers for readability
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import settings


# ── Style constants ───────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TOP_FILL     = PatternFill("solid", fgColor="E2EFDA")   # top-10 light green
NORMAL_FILL  = PatternFill("solid", fgColor="FFFFFF")
ALT_FILL     = PatternFill("solid", fgColor="F5F5F5")
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)

COLUMNS = [
    ("Rank",              6),
    ("Name",             22),
    ("Email",            28),
    ("Phone",            14),
    ("Skills",           40),
    ("Experience (yrs)", 14),
    ("Semantic /100",    14),
    ("Skill /100",       12),
    ("Exp /100",         10),
    ("Final Score /100", 14),
    ("Resume File",      30),
]

# Column index (1-based) for file_hash — stored hidden for dedup checks
HASH_COL = len(COLUMNS) + 1   # appended after visible columns, hidden


# ── Helpers ───────────────────────────────────────────────────────────────────

def _pct(value) -> int:
    """Convert 0.0–1.0 float to 0–100 int."""
    try:
        return round(float(value or 0) * 100)
    except (TypeError, ValueError):
        return 0


def _safe_sheet_name(role_name: str) -> str:
    """Excel sheet names: max 31 chars, no special chars."""
    return role_name[:31].replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "").replace("[", "").replace("]", "")


def _write_header_row(ws) -> None:
    """Write the frozen header row with styles."""
    for col_idx, (col_name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Hidden hash column for dedup
    ws.cell(row=1, column=HASH_COL, value="_file_hash")
    ws.column_dimensions[get_column_letter(HASH_COL)].hidden = True
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"


def _get_existing_hashes(ws) -> set[str]:
    """Read all file_hashes already in a sheet (for dedup)."""
    hashes = set()
    for row in ws.iter_rows(min_row=2, min_col=HASH_COL, max_col=HASH_COL, values_only=True):
        if row[0]:
            hashes.add(str(row[0]))
    return hashes


def _candidate_to_row(c: dict) -> list:
    skills_str = ", ".join(c.get("skills") or [])[:200]
    return [
        c.get("rank", ""),
        c.get("name", ""),
        c.get("email", ""),
        c.get("phone", ""),
        skills_str,
        c.get("experience_years", ""),
        _pct(c.get("semantic_score")),
        _pct(c.get("skill_score")),
        _pct(c.get("experience_score")),
        _pct(c.get("final_score")),
        c.get("file_name", ""),
        c.get("file_hash", ""),   # hidden dedup column
    ]


def _append_candidates_to_sheet(ws, new_candidates: list[dict]) -> int:
    """
    Append only candidates not already in the sheet.
    Returns count of rows actually added.
    """
    existing_hashes = _get_existing_hashes(ws)
    added = 0

    for cand in new_candidates:
        file_hash = cand.get("file_hash", "")
        if file_hash and file_hash in existing_hashes:
            logger.debug(f"Excel: skipping duplicate {cand.get('file_name')} (hash already in sheet)")
            continue

        next_row = ws.max_row + 1
        row_data = _candidate_to_row(cand)
        rank = cand.get("rank", 99)
        is_top = isinstance(rank, int) and rank <= 10

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=value)
            cell.fill = TOP_FILL if is_top else (NORMAL_FILL if next_row % 2 == 0 else ALT_FILL)
            cell.alignment = CENTER if col_idx in (1, 6, 7, 8, 9, 10) else LEFT

        existing_hashes.add(file_hash)
        added += 1

    return added


def _re_rank_sheet(ws) -> None:
    """
    After appending, re-sort all data rows by Final Score /100 descending
    and reassign rank numbers.
    """
    if ws.max_row < 3:
        return

    # Read all data rows (skip header row 1)
    data_rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if any(cell is not None for cell in row):
            data_rows.append(list(row))

    if not data_rows:
        return

    # Sort by Final Score /100 (column index 9, 0-based) descending
    SCORE_IDX = 9   # 0-based index of "Final Score /100" column
    data_rows.sort(key=lambda r: (r[SCORE_IDX] or 0), reverse=True)

    # Reassign ranks and rewrite rows
    current_rank = 1
    prev_score = None
    for i, row_data in enumerate(data_rows):
        score = row_data[SCORE_IDX] or 0
        if prev_score is not None and score < prev_score:
            current_rank = i + 1
        row_data[0] = current_rank   # update rank column
        prev_score = score

        excel_row = i + 2   # +2 because row 1 is header
        is_top = current_rank <= 10
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.fill = TOP_FILL if is_top else (NORMAL_FILL if excel_row % 2 == 0 else ALT_FILL)
            cell.alignment = CENTER if col_idx in (1, 6, 7, 8, 9, 10) else LEFT


def _write_summary_sheet(wb: Workbook, role_stats: dict[str, int]) -> None:
    """Create or overwrite the Summary sheet."""
    if "Summary" in wb.sheetnames:
        del wb["Summary"]

    ws = wb.create_sheet("Summary", 0)
    ws["A1"] = "ATS Candidate Ranking Report"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="1F3864")

    ws["A3"] = "Role"
    ws["B3"] = "Total Candidates"
    ws["C3"] = "Sheet"

    for cell in [ws["A3"], ws["B3"], ws["C3"]]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    for i, (role, count) in enumerate(sorted(role_stats.items()), start=4):
        ws.cell(row=i, column=1, value=role)
        ws.cell(row=i, column=2, value=count)
        ws.cell(row=i, column=3, value=_safe_sheet_name(role))

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18




# ── Top 5 per role sheet ──────────────────────────────────────────────────────

def _write_top5_sheet(wb: Workbook, role_candidates: dict[str, list[dict]]) -> None:
    """
    Creates a "Top 5 Candidates" sheet showing the best candidate
    per role with a hyperlink to the resume file.
    """
    if "Top 5 Candidates" in wb.sheetnames:
        del wb["Top 5 Candidates"]

    ws = wb.create_sheet("Top 5 Candidates", 1)

    # Title
    ws["A1"] = "Top 5 Candidates per Role"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="1F3864")
    ws["A2"] = "Candidates ranked by Final Score (out of 100)"
    ws["A2"].font = Font(italic=True, size=10, color="595959")

    TOP5_COLS = [
        ("Rank",              5),
        ("Role",             20),
        ("Name",             22),
        ("Email",            28),
        ("Phone",            14),
        ("Experience (yrs)", 12),
        ("Final Score /100", 14),
        ("Semantic /100",    13),
        ("Skill /100",       11),
        ("Skills (top 5)",   35),
        ("Resume",           20),
    ]

    header_row = 4
    for col_idx, (col_name, width) in enumerate(TOP5_COLS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[header_row].height = 22
    ws.freeze_panes = f"A{header_row + 1}"

    current_row = header_row + 1
    GREEN_FONT  = Font(name="Calibri", color="0563C1", underline="single", size=11)

    for role_name in sorted(role_candidates.keys()):
        candidates = role_candidates[role_name]
        if not candidates:
            continue

        # Sort and take top 5
        top5 = sorted(candidates, key=lambda c: c.get("final_score") or 0, reverse=True)[:5]

        for rank_in_role, cand in enumerate(top5, start=1):
            is_top = rank_in_role == 1
            fill = TOP_FILL if is_top else (NORMAL_FILL if current_row % 2 == 0 else ALT_FILL)

            skills_preview = ", ".join((cand.get("skills") or [])[:5])
            file_path = cand.get("file_path", "")
            file_name = cand.get("file_name", "")

            row_values = [
                rank_in_role,
                role_name,
                cand.get("name", ""),
                cand.get("email", ""),
                cand.get("phone", ""),
                cand.get("experience_years", ""),
                _pct(cand.get("final_score")),
                _pct(cand.get("semantic_score")),
                _pct(cand.get("skill_score")),
                skills_preview,
                file_name,   # hyperlink cell
            ]

            for col_idx, value in enumerate(row_values, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.fill = fill
                cell.alignment = CENTER if col_idx in (1, 6, 7, 8, 9) else LEFT

                # Make resume filename a hyperlink if file exists
                if col_idx == 11 and file_path and Path(file_path).exists():
                    cell.hyperlink = f"file:///{file_path.replace(chr(92), '/')}"
                    cell.font = GREEN_FONT
                    cell.value = file_name

            current_row += 1

        # Add a blank separator row between roles
        current_row += 1

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(TOP5_COLS))}{header_row}"


# ── Main public function ──────────────────────────────────────────────────────
def generate_excel_report(
    role_candidates: dict[str, list[dict]],
    output_path: Optional[Path] = None,
) -> Path:
    """
    Append new candidates to the Excel report.

    role_candidates: {role_name: [candidate_dict, ...]}

    Behaviour:
    - If file doesn't exist → create fresh
    - If file exists:
        - Sheet exists for role → append only new candidates (dedup by file_hash)
        - Sheet missing for role → create it
    - After all appends → re-rank each updated sheet
    - Always regenerate the Summary sheet
    """
    dest = output_path or settings.CANDIDATE_RANKING_OUTPUT
    dest.parent.mkdir(parents=True, exist_ok=True)

    # Load existing workbook or create fresh one
    if dest.exists():
        try:
            wb = load_workbook(str(dest))
            logger.info(f"Excel: loaded existing file {dest}")
        except Exception as e:
            logger.warning(f"Excel: could not load existing file ({e}), creating fresh")
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        logger.info(f"Excel: creating new file {dest}")

    role_stats: dict[str, int] = {}

    # Collect existing role stats from sheets already in the file
    for sheet_name in wb.sheetnames:
        if sheet_name == "Summary":
            continue
        ws = wb[sheet_name]
        # Count data rows (excluding header)
        count = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in row))
        role_stats[sheet_name] = count

    # Process each role
    for role_name, candidates in sorted(role_candidates.items()):
        if not candidates:
            continue

        safe_name = _safe_sheet_name(role_name)

        if safe_name in wb.sheetnames:
            # Sheet exists — append only new candidates
            ws = wb[safe_name]
            added = _append_candidates_to_sheet(ws, candidates)
            logger.info(f"Excel sheet '{safe_name}': {added} new candidate(s) added (dedup applied)")
        else:
            # New sheet — create with header and all candidates
            ws = wb.create_sheet(title=safe_name)
            _write_header_row(ws)
            added = _append_candidates_to_sheet(ws, candidates)
            logger.info(f"Excel sheet '{safe_name}': created with {added} candidate(s)")

        # Re-sort and re-rank the sheet after appending
        _re_rank_sheet(ws)

        # Update stats
        count = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in row))
        role_stats[safe_name] = count

    # Regenerate summary sheet
    _write_summary_sheet(wb, role_stats)

    # Generate Top 5 sheet
    _write_top5_sheet(wb, role_candidates)

    wb.save(str(dest))
    total = sum(role_stats.values())
    logger.info(f"Excel report saved → {dest} ({len(role_stats)} roles, {total} total candidates)")
    return dest